import os
import sys
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox, filedialog
import sqlite3
from collections import defaultdict
from datetime import datetime
from scipy.optimize import linprog
import threading
import time
from PIL import Image, ImageTk
import folium
import webbrowser
import requests

DATABASE_PATH = "//10.193.232.18/Java/우현 테스트/data_needed"
EXCLUDED_CARRIERS_DB = f"{DATABASE_PATH}/excluded_carriers.db"
AVAILABLE_TRUCKS_DB = f"{DATABASE_PATH}/available_trucks.db"
SHIPPING_POSTAL_CODES_DB = f"{DATABASE_PATH}/shipping_postal_codes.db"
CARRIER_ASSIGNMENT_DB = f"{DATABASE_PATH}/Carrier_assignment.db"

class DatabaseManager:
    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = None

    def connect(self):
        try:
            self.conn = sqlite3.connect(self.db_path)
            return self.conn
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to connect to database: {e}")
            return None

    def close(self):
        if self.conn:
            self.conn.close()

    def execute_query(self, query, params=()):
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, params)
            self.conn.commit()
            return cursor
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to execute query: {e}")
            return None

class ShippingCalculator:
    def __init__(self, root):
        self.root = root
        self.root.title("LX TMS")

        self.style = ttk.Style()
        self.style.theme_use('superhero')

        self.setup_start_interface()

        self.excluded_carriers = set()
        self.input_destination = {}  # 입력값을 관리하는 딕셔너리
        self.destination_entries = []
        self.truck_entries = []
        self.truck_type_vars = []
        self.truck_type_menus = []

        self.excluded_carriers_db = DatabaseManager(EXCLUDED_CARRIERS_DB)
        self.available_trucks_db = DatabaseManager(AVAILABLE_TRUCKS_DB)
        self.shipping_postal_codes_db = DatabaseManager(SHIPPING_POSTAL_CODES_DB)
        self.carrier_assignment_db = DatabaseManager(CARRIER_ASSIGNMENT_DB)

        # 하드 코딩
        #self.excluded_carriers_db = DatabaseManager("C:/Users/SISTEMAS/Documents/excluded_carriers.db")
        #self.available_trucks_db = DatabaseManager("C:/Users/SISTEMAS/Documents/available_trucks.db")
        #self.shipping_postal_codes_db = DatabaseManager("C:/Users/SISTEMAS/Documents/shipping_postal_codes.db")
        #self.carrier_assignment_db = DatabaseManager("C:/Users/SISTEMAS/Documents/Carrier_assignment.db")

        self.init_db()

    def init_db(self):
        self.excluded_carriers_db.connect()
        self.excluded_carriers_db.execute_query('''
            CREATE TABLE IF NOT EXISTS excluded_carriers (
                carrier TEXT,
                destination_postal_code INTEGER,
                PRIMARY KEY (carrier, destination_postal_code)
            )
        ''')
        self.excluded_carriers_db.close()

        # Carrier_assignment 테이블이 있는지 확인
        self.carrier_assignment_db.connect()
        self.carrier_assignment_db.execute_query('''
            CREATE TABLE IF NOT EXISTS Carrier_assignment (
                postal_code INTEGER,
                carrier TEXT,
                type TEXT,
                assigned_truck INTEGER
            )
        ''')
        self.carrier_assignment_db.close()

    def get_assigned_trucks(self):
        """Carrier_assignment 테이블에서 이미 할당된 트럭 수를 계산하는 함수"""
        assigned_trucks = defaultdict(lambda: defaultdict(int))  # carrier -> truck_type -> assigned_trucks
        self.carrier_assignment_db.connect()

        try:
            cursor = self.carrier_assignment_db.execute_query(
                'SELECT carrier, type, SUM(assigned_truck) FROM Carrier_assignment GROUP BY carrier, type')
            rows = cursor.fetchall()

            # 각 운송사 및 트럭 타입별로 할당된 트럭 수를 기록
            for row in rows:
                carrier = row[0]
                truck_type = row[1]
                total_assigned = row[2]
                assigned_trucks[carrier][truck_type] = total_assigned
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to retrieve assigned trucks: {e}")
        finally:
            self.carrier_assignment_db.close()

        return assigned_trucks

    def save_to_database(self, all_results, tree):
        """최적화된 결과를 데이터베이스에 저장하는 함수"""
        self.carrier_assignment_db.connect()
        self.available_trucks_db.connect()

        try:
            tree_children = tree.get_children()  # Treeview에 출력된 모든 항목을 가져옴
            if not tree_children:
                messagebox.showwarning("Warning", "No items to save.")
                return

            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 현재 날짜와 시간 가져오기

            for i, child in enumerate(tree_children):
                item_values = tree.item(child, 'values')  # 각 항목의 값 가져오기
                input_postal_code = item_values[0]  # Postal code
                carrier = item_values[1]  # Carrier
                truck_type = item_values[2]  # Truck type
                assigned_truck = int(item_values[3])  # Assigned truck

                # Carrier_assignment에 할당 정보 저장
                self.carrier_assignment_db.execute_query(
                    'INSERT INTO Carrier_assignment (postal_code, carrier, type, assigned_truck, Time) VALUES (?, ?, ?, ?, ?)',
                    (input_postal_code, carrier, truck_type, assigned_truck, current_time)
                )

                # Available_trucks에서 해당 운송사와 트럭 타입의 트럭 수를 업데이트
                cursor = self.available_trucks_db.execute_query(
                    'SELECT total_trucks FROM available_trucks WHERE carrier = ? AND truck_type = ?',
                    (carrier, truck_type)
                )
                result = cursor.fetchone()
                if result:
                    available_trucks = result[0]
                    new_total_trucks = available_trucks - assigned_truck

                    # 남은 트럭 수가 0 이상이어야 함
                    if new_total_trucks < 0:
                        messagebox.showerror("Error", f"Not enough trucks available for {carrier} ({truck_type})")
                        return

                    # 트럭 수 업데이트
                    self.available_trucks_db.execute_query(
                        'UPDATE available_trucks SET total_trucks = ? WHERE carrier = ? AND truck_type = ?',
                        (new_total_trucks, carrier, truck_type)
                    )

                # Treeview 항목에 태그를 추가하여 색상 변경 (저장 완료를 표시)
                tree.item(child, tags=("saved",))

            # 성공 메시지 표시
            messagebox.showinfo("Success", "All data has been successfully saved to the database.")

            # 저장된 항목들의 색상을 변경 (예: 녹색)
            tree.tag_configure('saved', foreground='green')

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to save data: {e}")
        finally:
            self.carrier_assignment_db.close()
            self.available_trucks_db.close()

    def reset_carrier_assignment_db(self):
        """Carrier_assignment.db의 데이터를 날짜별로 리셋하는 함수"""
        try:
            # Carrier_assignment 테이블에서 모든 날짜(Time)를 가져옵니다.
            self.carrier_assignment_db.connect()

            cursor = self.carrier_assignment_db.execute_query(
                'SELECT DISTINCT Time FROM Carrier_assignment ORDER BY Time')
            date_rows = cursor.fetchall()

            self.carrier_assignment_db.close()

            if not date_rows:
                messagebox.showwarning("No Data", "No data available to reset.")
                return

            # 날짜 선택을 위한 새 창을 엽니다.
            date_window = tk.Toplevel(self.root)
            date_window.title("Select Date to Reset")

            # 라벨 추가
            label = ttk.Label(date_window, text="Select a date to reset:", font=("impact", 12))
            label.pack(pady=10)

            # 날짜 선택 콤보박스
            selected_date = tk.StringVar()
            date_combo = ttk.Combobox(date_window, textvariable=selected_date, state="readonly", font=("Arial", 10))
            date_combo['values'] = [row[0] for row in date_rows]  # 각 날짜 값을 콤보박스에 추가
            date_combo.pack(pady=10)

            # 선택한 날짜의 데이터를 삭제하는 함수
            def delete_selected_date():
                selected = selected_date.get()

                if selected:
                    confirmation = messagebox.askyesno("Reset Confirmation",
                                                       f"Do you want to reset assignments for {selected}?")
                    if confirmation:
                        try:
                            self.carrier_assignment_db.connect()

                            # 선택한 날짜에 해당하는 할당된 트럭 수를 가져옵니다.
                            cursor = self.carrier_assignment_db.execute_query(
                                'SELECT carrier, type, SUM(assigned_truck) FROM Carrier_assignment WHERE Time = ? GROUP BY carrier, type',
                                (selected,)
                            )
                            assigned_trucks = cursor.fetchall()

                            # 해당 트럭 수를 available_trucks.db에 다시 더해줍니다.
                            self.available_trucks_db.connect()
                            for row in assigned_trucks:
                                carrier = row[0]
                                truck_type = row[1]
                                assigned_truck = row[2]

                                # 남은 트럭 수를 업데이트
                                cursor = self.available_trucks_db.execute_query(
                                    'SELECT total_trucks FROM available_trucks WHERE carrier = ? AND truck_type = ?',
                                    (carrier, truck_type)
                                )
                                result = cursor.fetchone()
                                if result:
                                    available_trucks = result[0]
                                    new_total_trucks = available_trucks + assigned_truck
                                    self.available_trucks_db.execute_query(
                                        'UPDATE available_trucks SET total_trucks = ? WHERE carrier = ? AND truck_type = ?',
                                        (new_total_trucks, carrier, truck_type)
                                    )

                            # 선택한 날짜의 데이터 삭제
                            self.carrier_assignment_db.execute_query(
                                'DELETE FROM Carrier_assignment WHERE Time = ?', (selected,))
                            messagebox.showinfo("Success", f"All assignments for {selected} have been reset.")
                            date_window.destroy()

                        except sqlite3.Error as e:
                            messagebox.showerror("Database Error", f"Failed to reset data: {e}")
                        finally:
                            self.carrier_assignment_db.close()
                            self.available_trucks_db.close()

                else:
                    messagebox.showwarning("Selection Error", "Please select a date.")

            # 리셋 버튼 추가
            reset_button = ttk.Button(date_window, text="Reset", command=delete_selected_date, bootstyle="danger")
            reset_button.pack(pady=20)

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to retrieve data: {e}")
        finally:
            if self.carrier_assignment_db.conn:
                self.carrier_assignment_db.close()

        date_window.geometry("300x200")

    def show_carrier_assignment_detail(self):
        """Carrier_assignment.db 파일 내용을 날짜별로 탭을 생성하여 보여주는 함수"""
        self.carrier_assignment_db.connect()

        # 새 창을 열어 결과를 표시
        detail_window = tk.Toplevel(self.root)
        detail_window.title("Carrier Assignment Details")

        notebook = ttk.Notebook(detail_window)  # Notebook 생성
        notebook.pack(expand=True, fill="both")

        try:
            # Carrier_assignment 테이블의 모든 데이터를 조회
            cursor = self.carrier_assignment_db.execute_query(
                'SELECT DISTINCT Time FROM Carrier_assignment ORDER BY Time')
            time_rows = cursor.fetchall()

            # 저장된 날짜별로 탭 생성
            for time_row in time_rows:
                time_value = time_row[0]
                tab_frame = ttk.Frame(notebook)  # 각 날짜별 Frame 생성
                notebook.add(tab_frame, text=time_value)  # 탭 추가 (시간을 탭 제목으로 설정)

                # Treeview 생성
                tree = ttk.Treeview(tab_frame, columns=("Postal Code", "Carrier", "Type", "Assigned Truck", "Time"),
                                    show="headings")
                tree.heading("Postal Code", text="Postal Code")
                tree.heading("Carrier", text="Carrier")
                tree.heading("Type", text="Type")
                tree.heading("Assigned Truck", text="Assigned Truck")
                tree.heading("Time", text="Assigned Time")

                tree.pack(expand=True, fill="both")

                # 선택된 Time 값에 해당하는 데이터를 가져와 Treeview에 표시
                cursor = self.carrier_assignment_db.execute_query(
                    'SELECT postal_code, carrier, type, assigned_truck, Time FROM Carrier_assignment WHERE Time = ?',
                    (time_value,)
                )
                rows = cursor.fetchall()

                for row in rows:
                    tree.insert("", "end", values=row)

                # 우클릭 이벤트 바인딩
                tree.bind("<Button-3>", lambda event, t=tree: self.on_carrier_assignment_right_click(event, t))

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to retrieve data from Carrier_assignment.db: {e}")
        finally:
            self.carrier_assignment_db.close()

        detail_window.geometry("800x400")

    def on_carrier_assignment_right_click(self, event, tree):
        """우클릭 이벤트 처리하여 컨텍스트 메뉴 생성 및 항목 삭제 기능 추가"""
        selected_item = tree.selection()

        if not selected_item:
            return

        selected_item = selected_item[0]
        values = tree.item(selected_item, "values")  # 선택된 항목의 값을 가져옴
        postal_code = values[0]
        carrier = values[1]

        # 컨텍스트 메뉴 생성
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Delete",
                         command=lambda: self.delete_carrier_assignment(tree, selected_item, postal_code, carrier))
        menu.post(event.x_root, event.y_root)

    def delete_carrier_assignment(self, tree, selected_item, postal_code, carrier):
        """Carrier_assignment.db에서 선택된 항목을 삭제하는 함수"""
        confirmation = messagebox.askyesno("Delete Confirmation",
                                           f"Do you really want to delete the assignment for {carrier} at {postal_code}?")

        if confirmation:
            try:
                self.carrier_assignment_db.connect()

                # 삭제할 트럭 수를 가져옵니다.
                cursor = self.carrier_assignment_db.execute_query(
                    'SELECT carrier, type, assigned_truck FROM Carrier_assignment WHERE postal_code = ? AND carrier = ?',
                    (postal_code, carrier)
                )
                assignment = cursor.fetchone()

                if assignment:
                    truck_type = assignment[1]
                    assigned_truck = assignment[2]

                    # available_trucks.db에 트럭 수를 복구합니다.
                    self.available_trucks_db.connect()
                    cursor = self.available_trucks_db.execute_query(
                        'SELECT total_trucks FROM available_trucks WHERE carrier = ? AND truck_type = ?',
                        (carrier, truck_type)
                    )
                    result = cursor.fetchone()
                    if result:
                        available_trucks = result[0]
                        new_total_trucks = available_trucks + assigned_truck
                        self.available_trucks_db.execute_query(
                            'UPDATE available_trucks SET total_trucks = ? WHERE carrier = ? AND truck_type = ?',
                            (new_total_trucks, carrier, truck_type)
                        )

                    # Carrier_assignment.db에서 할당을 삭제합니다.
                    self.carrier_assignment_db.execute_query(
                        'DELETE FROM Carrier_assignment WHERE postal_code = ? AND carrier = ?',
                        (postal_code, carrier)
                    )
                    messagebox.showinfo("Success", "The assignment has been deleted.")
                    tree.delete(selected_item)

            except sqlite3.Error as e:
                messagebox.showerror("Database Error", f"Failed to delete the assignment: {e}")
            finally:
                self.carrier_assignment_db.close()
                self.available_trucks_db.close()

    def get_excluded_carriers(self):
        """데이터베이스에서 제외된 운송사 목록을 가져오는 함수"""
        self.excluded_carriers_db.connect()
        cursor = self.excluded_carriers_db.execute_query(
            'SELECT carrier, destination_postal_code FROM excluded_carriers'
        )
        carriers = {(row[0], row[1]) for row in cursor.fetchall()} if cursor else set()
        self.excluded_carriers_db.close()
        return carriers

    def add_to_excluded_carriers(self, carrier, input_postal_code):
        """특정 input postal code의 운송사를 제외 목록에 추가"""
        self.excluded_carriers_db.connect()
        self.excluded_carriers_db.execute_query('''
            INSERT OR IGNORE INTO excluded_carriers (carrier, destination_postal_code) 
            VALUES (?, ?)
        ''', (carrier, input_postal_code))  # input_postal_code만 추가
        self.excluded_carriers_db.close()

    def remove_from_excluded_carriers(self, carrier, destination_postal_code):
        """특정 우편번호의 운송사를 제외 목록에서 제거"""
        self.excluded_carriers_db.connect()
        try:
            self.excluded_carriers_db.execute_query(
                'DELETE FROM excluded_carriers WHERE carrier = ? AND destination_postal_code = ?',
                (carrier, destination_postal_code)
            )
            self.excluded_carriers_db.conn.commit()

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to remove carrier: {e}")
        finally:
            self.excluded_carriers_db.close()

    def clear_excluded_carriers(self):
        """제외된 운송사 목록을 초기화"""
        self.excluded_carriers_db.connect()
        self.excluded_carriers_db.execute_query('DELETE FROM excluded_carriers')
        self.excluded_carriers_db.close()

    def fetch_shipping_rates(self, destination, truck_types):
        """해당 목적지와 트럭 타입에 대한 운송사 정보를 가져오는 함수"""
        self.shipping_postal_codes_db.connect()
        data = []
        try:
            cursor = self.shipping_postal_codes_db.conn.cursor()
            for truck_type in truck_types:
                cursor.execute(
                    '''SELECT carrier, vehicle_type, origin, start_postal_code, end_postal_code, cost 
                       FROM shipping_postal_codes 
                       WHERE ? BETWEEN start_postal_code AND end_postal_code AND vehicle_type = ?''',
                    (destination, truck_type)
                )
                data.extend(cursor.fetchall())

            # 특정 input postal code에 대한 운송사만 제외 (범위 제외 대신)
            data = [row for row in data if (row[0], destination) not in self.excluded_carriers]
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to retrieve shipping rates: {e}")
            return None
        finally:
            self.shipping_postal_codes_db.close()

        # 중복된 carrier, vehicle_type, cost 조합을 제거
        unique_data = set()  # 중복 확인을 위한 세트
        filtered_data = []

        for row in data:
            key = (row[0], row[1], row[5])  # carrier, vehicle_type, cost 조합으로 중복 체크
            if key not in unique_data:
                unique_data.add(key)  # 중복되지 않으면 추가
                filtered_data.append(row)  # 데이터 저장

        return filtered_data

    def fetch_carrier_limits(self, carriers, truck_types):
        """운송사와 트럭 타입별 최대 배차 가능 수를 계산하는 함수"""
        carrier_limits = defaultdict(lambda: defaultdict(int))  # carrier -> truck_type -> limit
        assigned_trucks = self.get_assigned_trucks()  # 이미 할당된 트럭 수를 가져옴
        self.available_trucks_db.connect()

        try:
            cursor = self.available_trucks_db.conn.cursor()
            query = '''
                SELECT carrier, truck_type, SUM(total_trucks) FROM available_trucks 
                WHERE carrier IN ({carrier_seq}) AND truck_type IN ({truck_type_seq})
                GROUP BY carrier, truck_type
            '''
            # carriers와 truck_types를 각각의 쿼리에 맞게 placeholders로 설정
            query = query.format(
                carrier_seq=','.join('?' * len(carriers)),
                truck_type_seq=','.join('?' * len(truck_types))
            )
            cursor.execute(query, carriers + truck_types)

            for carrier, truck_type, limit in cursor.fetchall():
                # 사용 가능한 트럭 수에서 이미 할당된 트럭 수를 차감
                available_limit = limit - assigned_trucks.get((carrier, truck_type), 0)
                carrier_limits[carrier][truck_type] = max(available_limit, 0)  # 음수가 되지 않도록 함

        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to retrieve carrier limits: {e}")
        finally:
            self.available_trucks_db.close()

        return carrier_limits

    def calculate_optimal_shipping(self):
        try:
            raw_input_destination = [
                {
                    'postal_code': int(entry.get().strip()),
                    'trucks': int(self.truck_entries[i][1].get().strip() or 0),
                    'truck_type': self.truck_type_vars[i].get()
                }
                for i, (label, entry) in enumerate(self.destination_entries)
                if entry.get().strip()
            ]

            # 동일한 postal code와 truck type을 합산
            combined_routes = defaultdict(lambda: {'trucks': 0, 'postal_code': None, 'truck_type': None})

            for entry in raw_input_destination:
                key = (entry['postal_code'], entry['truck_type'])  # postal code와 truck type을 키로 사용
                combined_routes[key]['trucks'] += entry['trucks']  # 트럭 수 합산
                combined_routes[key]['postal_code'] = entry['postal_code']
                combined_routes[key]['truck_type'] = entry['truck_type']

            self.input_destination = list(combined_routes.values())

            if not self.input_destination:
                messagebox.showerror("Error",
                                     "Please enter all fields including destination, truck count, and truck type.")
                return

            self.excluded_carriers.update(self.get_excluded_carriers())

            # 모든 입력 데이터를 통합하여 처리할 수 있도록 데이터 모으기
            all_data = []
            truck_requirements = []
            route_ids = []

            for entry in self.input_destination:
                destination = entry['postal_code']
                truck_type = entry['truck_type']
                required_trucks = entry['trucks']

                # 각 입력에 대한 데이터를 가져옴
                data = self.fetch_shipping_rates(destination, [truck_type])
                if not data:
                    messagebox.showerror("Error",
                                         f"No data found for the postal code {destination} and truck type {truck_type}.")
                    continue

                for d in data:
                    # input_postal_code와 트럭 타입을 각 데이터에 추가
                    d = list(d) + [destination, truck_type]
                    all_data.append(d)

                truck_requirements.append(required_trucks)
                route_ids.append((destination, truck_type))  # 고유한 루트 식별자

            # 최적화 과정은 전체 데이터를 대상으로 한 번에 실행
            result = self.optimize_shipping(all_data, truck_requirements, route_ids)

            if result:
                # 결과를 각 입력값에 대해 구분하여 출력
                self.display_results(result)
                self.all_results = result # 최적화 결과를 클래스 변수로 저장
            else:
                messagebox.showerror("Optimization Error", "Failed to find an optimal solution.")

        except ValueError as e:
            messagebox.showerror("Input Error", f"Invalid input: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")

    def optimize_shipping(self, all_data, truck_requirements, route_ids):
        try:
            c = [row[5] for row in all_data]  # 운임 (cost) 벡터
            carriers = [row[0] for row in all_data]  # 운송사
            vehicle_types = [row[1] for row in all_data]  # 트럭 타입
            input_postal_codes = [row[6] for row in all_data]  # 각 데이터의 목적지
            input_truck_types = [row[7] for row in all_data]  # 각 데이터의 트럭 타입
            num_data = len(all_data)

            # 각 루트별 제약 조건 행렬 설정
            A_eq = []
            b_eq = []

            for i, required_trucks in enumerate(truck_requirements):
                route_postal_code, route_truck_type = route_ids[i]
                row = [0] * num_data
                for j in range(num_data):
                    if input_postal_codes[j] == route_postal_code and input_truck_types[j] == route_truck_type:
                        row[j] = 1
                A_eq.append(row)
                b_eq.append(required_trucks)

            # 운송사 및 트럭 타입별로 제한을 설정
            carrier_limits = self.fetch_carrier_limits(carriers, vehicle_types)
            A_ub, b_ub = [], []
            if carrier_limits:
                for i, (carrier, truck_type) in enumerate(zip(carriers, vehicle_types)):
                    limit = carrier_limits.get(carrier, {}).get(truck_type, 0)
                    row = [1 if carriers[j] == carrier and vehicle_types[j] == truck_type else 0 for j in
                           range(num_data)]
                    A_ub.append(row)
                    b_ub.append(limit)

            bounds = [(0, None) for _ in range(num_data)]

            result = linprog(c, A_eq=A_eq, b_eq=b_eq, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method='highs')

            if result.success:
                return {
                    'carriers': carriers,
                    'vehicle_types': vehicle_types,
                    'input_postal_codes': input_postal_codes,
                    'input_truck_types': input_truck_types,
                    'assignments': result.x,
                    'total_cost': result.fun,
                    'costs': c
                }
            else:
                messagebox.showerror("Optimization Failed", "Optimization failed.")
                return None

        except Exception as e:
            messagebox.showerror("Optimization Error", f"An error occurred: {e}")
            return None

    def display_results(self, result):
        # 기존에 표시된 결과 삭제
        for widget in self.result_frame.winfo_children():
            widget.destroy()

        tree = ttk.Treeview(self.result_frame,
                            columns=("Input Postal Code", "Carrier", "Truck Type", "Assigned Trucks", "Cost"),
                            show="headings")

        tree.column("Input Postal Code", anchor="center", width=160, minwidth=160, stretch=False)
        tree.column("Carrier", anchor="center", width=160, minwidth=160, stretch=False)
        tree.column("Truck Type", anchor="center", width=160, minwidth=160, stretch=False)
        tree.column("Assigned Trucks", anchor="center", width=160, minwidth=160, stretch=False)
        tree.column("Cost", anchor="center", width=160, minwidth=160, stretch=False)

        tree.heading("Input Postal Code", text="Input Postal Code")
        tree.heading("Carrier", text="Carrier")
        tree.heading("Truck Type", text="Truck Type")
        tree.heading("Assigned Trucks", text="Assigned Trucks")
        tree.heading("Cost", text="Cost")

        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(self.result_frame, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(expand=True, fill="both")

        total_cost = 0

        for i, x in enumerate(result['assignments']):
            if x > 0:
                input_postal_code = result['input_postal_codes'][i]
                input_truck_type = result['input_truck_types'][i]
                tree.insert("", "end", values=(
                    input_postal_code, result['carriers'][i], input_truck_type, int(x),
                    f"${result['costs'][i] * x:,.2f}"))
                total_cost += result['costs'][i] * x

        total_cost_label = ttk.Label(self.result_frame, text=f"Total cost: ${total_cost:,.2f} (MXN)",
                                     font=('Arial', 15, 'bold'))
        total_cost_label.pack(pady=10)

        tree.bind("<Button-3>", lambda event: self.on_item_right_click(event, tree))

        # 데이터베이스 저장 버튼 추가
        db_save_button = ttk.Button(self.main_frame, text="Save",
                                    command=lambda: self.save_to_database(self.all_results, tree))

        db_save_button.grid(row=1, column=2, sticky="e", padx=10, pady=10)

        excluded_list_button = ttk.Button(self.main_frame, text="Exception", command=self.show_excluded_list,
                                          bootstyle="warning")
        excluded_list_button.grid(row=1, column=3, sticky="e", padx=10, pady=10)

    def on_item_right_click(self, event, tree):
        selected_items = tree.selection()

        if not selected_items:
            return

        selected_item = selected_items[0]
        values = tree.item(selected_item, "values")
        selected_carrier = values[1]  # Carrier 값을 가져옴
        input_postal_code = int(values[0])  # Input Postal Code 값을 가져옴
        truck_type = values[2]  # Truck Type 값을 가져옴

        try:
            self.show_context_menu(event, selected_carrier, input_postal_code, truck_type)
        except IndexError as e:
            messagebox.showerror("Error", f"Failed to parse route information: {str(e)}")

    def show_context_menu(self, event, carrier, input_postal_code, truck_type):
        """컨텍스트 메뉴를 생성하여 운송사를 제외할 수 있게 하며, Detail 기능을 추가합니다."""
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Delete", command=lambda: self.exclude_carrier(carrier, input_postal_code))
        menu.add_command(label="Others", command=lambda: self.show_carrier_details(input_postal_code, truck_type))
        menu.post(event.x_root, event.y_root)

    def show_carrier_details(self, input_postal_code, truck_type):
        """Detail 버튼을 클릭하면 새 창에서 운송사 목록을 보여주는 함수"""
        details_window = tk.Toplevel(self.root)
        details_window.title(f"Details for {input_postal_code} ({truck_type})")

        # Treeview 생성
        tree = ttk.Treeview(details_window,
                            columns=("Carrier", "Truck Type", "Cost", "Available Limit", "Input Postal Code"),
                            show="headings")
        tree.heading("Carrier", text="Carrier")
        tree.heading("Truck Type", text="Truck Type")  # 트럭 타입 열 추가
        tree.heading("Cost", text="Cost")
        tree.heading("Available Limit", text="Available Limit")  # 사용 가능한 트럭 수 열 추가
        tree.heading("Input Postal Code", text="Input Postal Code")

        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(details_window, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        tree.pack(expand=True, fill="both")

        # 주어진 postal_code와 truck_type에 해당하는 운송사 데이터를 가져옴
        shipping_data = self.fetch_shipping_rates(input_postal_code, [truck_type])

        # 운송사 목록과 트럭 타입 목록을 모두 가져와 fetch_carrier_limits에 전달
        carriers = [row[0] for row in shipping_data]
        truck_types = [row[1] for row in shipping_data]

        # 운송사 및 트럭 타입별로 사용 가능한 트럭 수를 가져옴
        carrier_limits = self.fetch_carrier_limits(carriers, truck_types)  # 운송사별 사용 가능한 트럭 수

        # 운임 비용 기준으로 정렬
        sorted_data = sorted(shipping_data, key=lambda x: x[5])

        for row in sorted_data:
            carrier = row[0]
            truck_type = row[1]  # 트럭 타입 추가
            cost = row[5]
            available_limit = carrier_limits.get(carrier, {}).get(truck_type, 0)  # 해당 운송사와 트럭 타입의 사용 가능한 트럭 수를 가져옴
            tree.insert("", "end", values=(carrier, truck_type, f"${cost:,.2f}", available_limit, input_postal_code))

        details_window.geometry("1000x400")

    def exclude_carrier(self, carrier, input_postal_code):
        """특정 input postal code에 대한 운송사를 제외 목록에 추가하고 알림을 표시합니다."""
        # 이미 존재하는지 확인
        existing_exclusions = self.get_excluded_carriers()

        # 특정 input postal code만 제외
        if (carrier, input_postal_code) not in existing_exclusions:
            self.excluded_carriers.add((carrier, input_postal_code))
            self.add_to_excluded_carriers(carrier, input_postal_code)  # 범위 대신 input postal code만 추가
            messagebox.showinfo("Excluded", f"{carrier} has been excluded for postal code {input_postal_code}.")
        else:
            messagebox.showinfo("Already Excluded",
                                f"{carrier} is already excluded for postal code {input_postal_code}.")

    def show_excluded_list(self):
        """제외된 운송사 목록을 보여주는 창 생성"""
        excluded_window = tk.Toplevel(self.root)
        excluded_window.title("List of exception carriers")

        # Treeview를 생성할 때 'carrier'와 'destination_postal_code'가 열 이름으로 설정됩니다.
        tree = ttk.Treeview(excluded_window, columns=("carrier", "destination_postal_code"), show="headings")
        tree.heading("carrier", text="Excluded carriers")
        tree.heading("destination_postal_code", text="Postal Code")

        excluded_list = self.get_excluded_carriers()  # 제외된 운송사 목록 가져오기
        for carrier, destination_postal_code in excluded_list:
            tree.insert("", "end", values=(carrier, destination_postal_code))

        tree.pack(expand=True, fill="both")

        # 리셋 버튼 추가
        reset_button = ttk.Button(excluded_window, text="Reset", command=lambda: self.reset_excluded_list(tree),
                                  bootstyle="danger")
        reset_button.pack(pady=10)

        # Treeview에 우클릭 이벤트 추가
        tree.bind("<Button-3>", lambda event: self.on_excluded_item_right_click(event, tree))

    def on_excluded_item_right_click(self, event, tree):
        selected_item = tree.selection()[0]
        values = tree.item(selected_item, "values")

        if len(values) < 2:
            messagebox.showerror("Error", "Selected item does not have carrier and postal code information.")
            return

        selected_carrier = values[0]
        selected_postal_code = values[1]

        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Return", command=lambda: self.restore_carrier(tree, selected_item, selected_carrier,
                                                                              selected_postal_code))
        menu.post(event.x_root, event.y_root)

    def restore_carrier(self, tree, item, carrier, postal_code):
        try:
            # 데이터베이스에서 항목 제거
            self.remove_from_excluded_carriers(carrier, postal_code)

            # 데이터베이스에서 항목이 제대로 제거되었는지 확인
            self.excluded_carriers_db.connect()
            cursor = self.excluded_carriers_db.execute_query(
                'SELECT * FROM excluded_carriers WHERE carrier = ? AND destination_postal_code = ?',
                (carrier, postal_code)
            )
            result = cursor.fetchall()
            self.excluded_carriers_db.close()

            if not result:
                self.excluded_carriers = self.get_excluded_carriers()
                self.refresh_treeview(tree)
                messagebox.showinfo("Returned", f"{carrier} has been returned for postal code {postal_code}.")
            else:
                messagebox.showerror("Error",
                                     f"Failed to remove {carrier} for postal code {postal_code} from the database.")

        except KeyError:
            messagebox.showerror("Error",
                                 f"Carrier {carrier} for postal code {postal_code} is not in the excluded list.")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    def refresh_treeview(self, tree):
        """Treeview를 갱신하는 함수"""
        for item in tree.get_children():
            tree.delete(item)

        excluded_list = self.get_excluded_carriers()
        for carrier, destination in excluded_list:
            tree.insert("", "end", values=(carrier, destination))

    def reset_excluded_list(self, tree):
        """제외 목록을 초기화하고 화면을 업데이트"""
        self.clear_excluded_carriers()
        self.excluded_carriers.clear()
        for item in tree.get_children():
            tree.delete(item)
        messagebox.showinfo("Reset", "Reset complete.")

    def save_results_to_file(self, all_results):
        """결과를 엑셀 파일로 저장하는 함수"""

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save the result file",
            initialfile=f"shipping_results_{datetime.now().strftime('%y-%m-%d %H')}.xlsx"
        )

        if not file_path:
            messagebox.showwarning("Warning", "No file selected. The file was not saved.")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Shipping Results"

            # 엑셀 헤더 추가
            ws.append(["Input Postal Code", "Carrier", "Truck Type", "Assigned Trucks", "Cost"])

            # 결과 데이터를 엑셀에 추가
            total_cost = 0
            if not isinstance(all_results, dict):
                messagebox.showerror("Error", "Invalid result structure, expected dictionary of results.")
                return

            # all_results가 딕셔너리이므로 바로 필드를 처리
            for i, x in enumerate(all_results['assignments']):
                if x > 0:
                    postal_code = all_results['input_postal_codes'][i]
                    carrier = all_results['carriers'][i]
                    truck_type = all_results['vehicle_types'][i]
                    assigned_trucks = int(x)
                    cost = all_results['costs'][i] * x

                    ws.append([postal_code, carrier, truck_type, assigned_trucks, f"{cost:,.2f}"])
                    total_cost += cost

            # 총 비용을 엑셀에 추가
            ws.append([])
            ws.append(["", "", "", "Total Cost", f"{total_cost:,.2f}"])

            # 엑셀 파일 저장
            wb.save(file_path)
            messagebox.showinfo("Save Complete", f"The results were saved at {file_path}.")

        except PermissionError:
            messagebox.showerror("Permission Error",
                                 f"Failed to save the file at {file_path}. Please check if the file is open or you have the necessary permissions.")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")

    def setup_start_interface(self):
        """기본 시작 인터페이스 설정"""
        self.start_frame = ttk.Frame(self.root)
        self.start_frame.pack(fill="both", expand=True)

        # 이미지 불러오기
        image_path = f"{DATABASE_PATH}/mexico_map.png"
        img = Image.open(image_path)
        img = img.resize((580, 580), Image.Resampling.LANCZOS)
        self.photo = ImageTk.PhotoImage(img)

        # Label을 사용하여 이미지 표시
        img_label = ttk.Label(self.start_frame, image=self.photo)
        img_label.pack(pady=20, side="bottom")

        # 텍스트 라벨 추가
        start_label = ttk.Label(self.start_frame, text="Transportation Management System", font=("impact", 24))
        start_label.pack(pady=50)

        # 로그인 버튼 추가
        login_button = ttk.Button(self.start_frame, text="Log in", command=self.show_main_screen, bootstyle="primary")
        login_button.pack(pady=20)

        # 설정 버튼 추가
        admin_button = ttk.Button(self.start_frame, text="Settings", command=self.show_admin_interface,
                                  bootstyle="secondary")
        admin_button.pack(pady=20)

    def show_main_screen(self):
        """메인 화면으로 이동"""
        self.start_frame.pack_forget()
        self.setup_main_interface()

    def show_admin_interface(self):
        """관리자 모드 화면"""
        self.start_frame.pack_forget()
        self.admin_frame = ttk.Frame(self.root)
        self.admin_frame.pack(fill="both", expand=True)

        admin_label = ttk.Label(self.admin_frame, text="Database", font=("impact", 18))
        admin_label.pack(pady=20)

        available_trucks_button = ttk.Button(self.admin_frame, text="Available_Truck", command=self.show_db_editor_interface, bootstyle="info")
        available_trucks_button.pack(pady=10)

        back_button = ttk.Button(self.admin_frame, text="Back", command=self.back_to_start, bootstyle="danger")
        back_button.pack(pady=10)

        self.root.geometry("500x500")

    def back_to_start(self):
        """시작 화면으로 돌아가는 함수"""
        self.admin_frame.pack_forget()
        self.start_frame.pack(fill="both", expand=True)

    def setup_main_interface(self):
        """메인 화면 설정"""
        if hasattr(self, 'main_frame'):
            self.main_frame.pack_forget()

        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill="both", expand=True)
        self.main_frame.grid_columnconfigure(0, weight=1, uniform="group1")
        self.main_frame.grid_columnconfigure(1, weight=1, uniform="group1")
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(1, weight=0)

        # 메뉴바 생성
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)

        # 엑셀 메뉴 추가
        excel_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Excel", menu=excel_menu)
        excel_menu.add_command(label="Download Excel Template", command=self.download_excel_template)
        excel_menu.add_command(label="Upload Excel File", command=self.upload_excel_file)

        excel_menu.add_command(
            label="Export Results to Excel",
            command=lambda: self.save_results_to_file(self.all_results) if hasattr(self,
                                                                                   'all_results') else messagebox.showwarning(
                "Warning", "No results to save.")
        )

        # Carrier assignment 메뉴 추가
        carrier_assignment_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Carrier assignment", menu=carrier_assignment_menu)

        # Carrier assignment 옵션 추가
        carrier_assignment_menu.add_command(label="Carrier assignment reset", command=self.reset_carrier_assignment_db)
        carrier_assignment_menu.add_command(label="Carrier assignment detail",
                                            command=self.show_carrier_assignment_detail)

        # 스크롤 가능 영역 생성
        input_frame_container = ttk.Frame(self.main_frame)
        input_frame_container.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # 캔버스 생성 및 스크롤바 연결
        canvas = tk.Canvas(input_frame_container)
        scrollbar = ttk.Scrollbar(input_frame_container, orient="vertical", command=canvas.yview)
        self.input_frame = ttk.Frame(canvas)

        self.input_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        self.root.geometry("1800x800")

        canvas.create_window((0, 0), window=self.input_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 캔버스와 스크롤바 레이아웃 설정
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # 결과 프레임 설정
        self.result_frame = ttk.Frame(self.main_frame)
        self.result_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        add_button = ttk.Button(self.input_frame, text="Add Postal code", command=self.add_input_fields,
                                bootstyle="success")
        add_button.grid(row=0, column=0, columnspan=2, pady=(0, 10))

        remove_button = ttk.Button(self.input_frame, text="Remove", command=self.remove_input_fields,
                                   bootstyle="danger")
        remove_button.grid(row=0, column=2, columnspan=2, pady=(0, 10))

        calculate_button = ttk.Button(self.input_frame, text="Calculate Optimal Cost",
                                      command=self.calculate_optimal_shipping, bootstyle="primary")
        calculate_button.grid(row=0, column=4, sticky="e", pady=(0, 10))

        back_button = ttk.Button(self.main_frame, text="Back", command=self.back_to_start_from_main, bootstyle="danger")
        back_button.place(relx=1.0, rely=0.0, anchor="ne", x=-10, y=10)

        self.destination_entries = []
        self.truck_entries = []
        self.truck_type_vars = []
        self.truck_type_menus = []
        self.add_input_fields()

        # 레이아웃 그리드 설정
        input_frame_container.grid_rowconfigure(0, weight=1)
        input_frame_container.grid_columnconfigure(0, weight=1)

    def back_to_start_from_main(self):
        """입력 페이지에서 시작 페이지로 돌아가는 함수"""
        self.main_frame.pack_forget()
        self.setup_start_interface()

    def add_input_fields(self):
        """입력 필드를 추가하는 함수"""
        row = len(self.destination_entries) + 1

        destination_label = ttk.Label(self.input_frame, text=f"Postal Code {row}")
        destination_label.grid(row=row, column=0)
        
        # postal code 입력을 위한 Combobox 생성
        destination_entry = ttk.Combobox(self.input_frame)
        destination_entry.grid(row=row, column=1, padx=5, pady=5, sticky="ew")
        
        # postal code 목록 가져오기
        postal_codes = self.fetch_postal_codes()
        destination_entry['values'] = postal_codes
        
        # 자동완성 이벤트 바인딩
        destination_entry.bind('<KeyRelease>', lambda event, entry=destination_entry: self.update_postal_code_suggestions(event, entry))

        trucks_label = ttk.Label(self.input_frame, text=f"Total Trucks {row}")
        trucks_label.grid(row=row, column=2)
        trucks_entry = ttk.Entry(self.input_frame)
        trucks_entry.grid(row=row, column=3, padx=5, pady=5, sticky="ew")

        truck_type_var = tk.StringVar()
        truck_type_menu = ttk.Combobox(self.input_frame, textvariable=truck_type_var)
        truck_type_menu.grid(row=row, column=4, padx=5, pady=5, sticky="ew")

        truck_types = self.fetch_truck_types()
        truck_type_menu['values'] = truck_types

        self.destination_entries.append((destination_label, destination_entry))
        self.truck_entries.append((trucks_label, trucks_entry))
        self.truck_type_vars.append(truck_type_var)
        self.truck_type_menus.append(truck_type_menu)

    def remove_input_fields(self):
        """가장 최근에 추가된 입력 필드를 제거하는 함수"""
        if self.destination_entries and self.truck_entries and self.truck_type_vars:
            destination_label, destination_entry = self.destination_entries.pop()
            trucks_label, trucks_entry = self.truck_entries.pop()
            truck_type_var = self.truck_type_vars.pop()
            truck_type_menu = self.truck_type_menus.pop()

            destination_label.grid_forget()
            destination_entry.grid_forget()
            trucks_label.grid_forget()
            trucks_entry.grid_forget()
            truck_type_menu.grid_forget()

    def fetch_truck_types(self):
        """데이터베이스에서 트럭 타입 목록을 가져오는 함수"""
        self.shipping_postal_codes_db.connect()
        cursor = self.shipping_postal_codes_db.execute_query('SELECT DISTINCT vehicle_type FROM shipping_postal_codes')
        truck_types = [row[0] for row in cursor.fetchall()] if cursor else []
        self.shipping_postal_codes_db.close()
        return truck_types

    def download_excel_template(self):
        """엑셀 템플릿을 다운로드하는 함수"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel Template",
            initialfile="shipping_template.xlsx"
        )

        if not file_path:
            messagebox.showwarning("Warning", "No file selected. The template was not saved.")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"

            ws.append(["Postal Code", "Total Trucks", "Truck Type"])  # 엑셀 템플릿의 헤더

            wb.save(file_path)
            messagebox.showinfo("Template Downloaded", f"The template was saved at {file_path}.")

        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    def upload_excel_file(self):
        """엑셀 파일을 업로드하여 데이터를 입력하는 함수"""

        def process_file(file_path):
            try:
                wb = load_workbook(filename=file_path)
                ws = wb.active

                # 템플릿 구조를 검증
                headers = [cell.value for cell in ws[1]]
                expected_headers = ["Postal Code", "Total Trucks", "Truck Type"]
                if headers != expected_headers:
                    messagebox.showerror("Error", "The selected file does not match the expected template format.")
                    return

                # 기존 입력 필드를 제거
                while self.destination_entries:
                    self.remove_input_fields()

                # 엑셀 데이터에 따라 새로운 입력 필드를 추가
                for row in ws.iter_rows(min_row=2, values_only=True):
                    postal_code, total_trucks, truck_type = row

                    if not all([postal_code, total_trucks, truck_type]):
                        continue  # 비어 있는 데이터는 건너뜀

                    self.add_input_fields()

                    self.destination_entries[-1][1].insert(0, str(postal_code))
                    self.truck_entries[-1][1].insert(0, str(total_trucks))
                    self.truck_type_vars[-1].set(truck_type)

            except Exception as e:
                messagebox.showerror("Error", f"An unexpected error occurred while processing the file: {str(e)}")
            finally:
                # 로딩 상태 종료
                self.stop_loading_animation()

        # 파일 선택 다이얼로그 열기
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Select Excel File"
        )

        if not file_path:
            messagebox.showwarning("Warning", "No file selected. The operation was cancelled.")
            return

        # 로딩 애니메이션 시작
        self.start_loading_animation()

        # 엑셀 파일 처리를 별도의 스레드에서 실행하여 UI가 멈추지 않도록 처리
        threading.Thread(target=process_file, args=(file_path,)).start()

    def start_loading_animation(self):
        """로딩 중임을 나타내는 애니메이션 시작"""
        self.loading_label = ttk.Label(self.main_frame, text="Uploading, please wait...",
                                       font=("Arial", 14))
        self.loading_label.grid(row=0, column=0, columnspan=2, pady=10)

        self.progress_bar = ttk.Progressbar(self.main_frame, mode='indeterminate')
        self.progress_bar.grid(row=1, column=0, columnspan=2, pady=10)
        self.progress_bar.start()

    def stop_loading_animation(self):
        """로딩 애니메이션을 중단"""
        self.progress_bar.stop()
        self.loading_label.grid_forget()  # 로딩 텍스트 제거
        self.progress_bar.grid_forget()  # 프로그레스 바 제거

    def show_db_editor_interface(self):
        """데이터베이스 편집 인터페이스를 설정하는 함수"""
        self.admin_frame.pack_forget()
        self.db_editor_frame = ttk.Frame(self.root)
        self.db_editor_frame.pack(fill="both", expand=True)

        db_label = ttk.Label(self.db_editor_frame, text="Available Trucks Editor", font=("impact", 14))
        db_label.pack(pady=10)

        # Treeview에 Truck Type 열 추가
        self.tree = ttk.Treeview(self.db_editor_frame, columns=("Carrier", "Truck Type", "Total Trucks"),
                                 show="headings")
        self.tree.heading("Carrier", text="carrier")
        self.tree.heading("Truck Type", text="Truck type")
        self.tree.heading("Total Trucks", text="Total truck")

        view_button = ttk.Button(self.db_editor_frame, text="View Data", command=self.view_data, bootstyle="info")
        view_button.pack(pady=5)

        add_button = ttk.Button(self.db_editor_frame, text="Add Entry", command=self.add_entry, bootstyle="success")
        add_button.pack(pady=5)

        edit_button = ttk.Button(self.db_editor_frame, text="Edit Entry", command=self.edit_entry, bootstyle="warning")
        edit_button.pack(pady=5)

        delete_button = ttk.Button(self.db_editor_frame, text="Delete Entry", command=self.delete_entry,
                                   bootstyle="danger")
        delete_button.pack(pady=5)

        back_button = ttk.Button(self.db_editor_frame, text="Back to Admin", command=self.back_to_admin,
                                 bootstyle="primary")
        back_button.pack(pady=5)

        self.tree.pack(expand=True, fill="both")

        self.root.geometry("1000x500")

    def back_to_admin(self):
        """관리자 모드로 돌아가는 함수"""
        self.db_editor_frame.pack_forget()
        self.admin_frame.pack(fill="both", expand=True)

    def view_data(self):
        """데이터베이스 데이터를 조회하는 함수"""
        self.available_trucks_db.connect()
        cursor = self.available_trucks_db.execute_query(
            'SELECT carrier, truck_type, total_trucks FROM available_trucks')
        rows = cursor.fetchall()
        self.available_trucks_db.close()

        for row in self.tree.get_children():
            self.tree.delete(row)

        for row in rows:
            self.tree.insert("", "end", values=row)

    def add_entry(self):
        """새로운 데이터를 추가하는 함수"""
        new_window = tk.Toplevel(self.root)
        new_window.title("Add Entry")

        carrier_label = ttk.Label(new_window, text="운송사")
        carrier_label.pack(pady=5)
        carrier_entry = ttk.Entry(new_window)
        carrier_entry.pack(pady=5)

        truck_type_label = ttk.Label(new_window, text="트럭 타입")
        truck_type_label.pack(pady=5)
        truck_type_entry = ttk.Entry(new_window)
        truck_type_entry.pack(pady=5)

        trucks_label = ttk.Label(new_window, text="총 차량 수")
        trucks_label.pack(pady=5)
        trucks_entry = ttk.Entry(new_window)
        trucks_entry.pack(pady=5)

        def save_new_entry():
            carrier = carrier_entry.get()
            truck_type = truck_type_entry.get()
            total_trucks = int(trucks_entry.get())
            self.available_trucks_db.connect()
            self.available_trucks_db.execute_query(
                'INSERT INTO available_trucks (carrier, truck_type, total_trucks) VALUES (?, ?, ?)',
                (carrier, truck_type, total_trucks)
            )
            self.available_trucks_db.close()
            self.view_data()
            new_window.destroy()

        save_button = ttk.Button(new_window, text="Save", command=save_new_entry, bootstyle="success")
        save_button.pack(pady=10)

    def edit_entry(self):
        selected_items = self.tree.selection()

        if not selected_items:
            messagebox.showwarning("Warning", "No item selected to edit.")
            return

        """선택된 데이터를 수정하는 함수"""
        selected_item = self.tree.selection()[0]
        selected_data = self.tree.item(selected_item, "values")

        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Entry")

        carrier_label = ttk.Label(edit_window, text="운송사")
        carrier_label.pack(pady=5)
        carrier_entry = ttk.Entry(edit_window)
        carrier_entry.pack(pady=5)
        carrier_entry.insert(0, selected_data[0])

        truck_type_label = ttk.Label(edit_window, text="트럭 타입")
        truck_type_label.pack(pady=5)
        truck_type_entry = ttk.Entry(edit_window)
        truck_type_entry.pack(pady=5)
        truck_type_entry.insert(0, selected_data[1])

        trucks_label = ttk.Label(edit_window, text="총 차량 수")
        trucks_label.pack(pady=5)
        trucks_entry = ttk.Entry(edit_window)
        trucks_entry.pack(pady=5)
        trucks_entry.insert(0, selected_data[2])

        def save_edit_entry():
            carrier = carrier_entry.get()
            truck_type = truck_type_entry.get()
            total_trucks = int(trucks_entry.get())
            self.available_trucks_db.connect()
            self.available_trucks_db.execute_query(
                'UPDATE available_trucks SET total_trucks = ?, truck_type = ? WHERE carrier = ? AND truck_type = ?',
                (total_trucks, truck_type, carrier, truck_type)
            )
            self.available_trucks_db.close()
            self.view_data()
            edit_window.destroy()

        save_button = ttk.Button(edit_window, text="Save", command=save_edit_entry, bootstyle="success")
        save_button.pack(pady=10)

    def delete_entry(self):
        selected_items = self.tree.selection()

        if not selected_items:
            messagebox.showwarning("Warning", "No item selected to delete.")
            return

        """선택된 데이터를 삭제하는 함수"""
        selected_item = self.tree.selection()[0]
        selected_data = self.tree.item(selected_item, "values")

        confirm = messagebox.askyesno("Delete Entry",
                                      f"Do you really want to delete {selected_data[0]} ({selected_data[1]})?")
        if confirm:
            self.available_trucks_db.connect()
            self.available_trucks_db.execute_query('DELETE FROM available_trucks WHERE carrier = ? AND truck_type = ?',
                                                   (selected_data[0], selected_data[1]))
            self.available_trucks_db.close()
            self.view_data()

    def fetch_postal_codes(self):
        """데이터베이스에서 사용 가능한 postal code 목록을 가져오는 함수"""
        self.shipping_postal_codes_db.connect()
        cursor = self.shipping_postal_codes_db.execute_query(
            'SELECT DISTINCT start_postal_code FROM shipping_postal_codes UNION SELECT DISTINCT end_postal_code FROM shipping_postal_codes'
        )
        postal_codes = sorted(list(set([str(row[0]) for row in cursor.fetchall()]))) if cursor else []
        self.shipping_postal_codes_db.close()
        return postal_codes

    def update_postal_code_suggestions(self, event, entry):
        """입력된 텍스트에 따라 postal code 제안을 업데이트하는 함수"""
        if not hasattr(self, '_postal_codes_cache'):
            self._postal_codes_cache = self.fetch_postal_codes()
        
        current_text = entry.get().strip()
        if current_text:
            suggestions = [code for code in self._postal_codes_cache if str(code).startswith(current_text)]
            entry['values'] = suggestions
        else:
            entry['values'] = self._postal_codes_cache


if __name__ == "__main__":
    root = ttk.Window(themename="superhero")
    app = ShippingCalculator(root)
    root.mainloop()

